from abc import ABC, abstractmethod
from openpyxl import Workbook, load_workbook


class Person(ABC):
    def __init__(self, name, phone):
        self.name = name
        self.phone = phone

    @abstractmethod
    def display_info(self):
        pass


class Customer(Person):
    discount_rate = 0.1  # Статическое свойство для всех клиентов

    def __init__(self, name, phone, loyalty_points=0):
        super().__init__(name, phone)
        self.loyalty_points = loyalty_points

    def display_info(self):
        print(f"Клиент: {self.name} - Телефон: {self.phone} - Баллы лояльности: {self.loyalty_points}")

    @staticmethod
    def set_discount_rate(rate):
        Customer.discount_rate = rate

    def earn_loyalty_points(self, points):
        self.loyalty_points += points

    def apply_discount(self, amount):
        return amount * (1 - Customer.discount_rate)


class BeautyService:
    def __init__(self, name, price):
        self.name = name
        self.price = price

    def display_info(self):
        print(f"Услуга: {self.name} - Цена: ${self.price}")


class BeautySalon:
    def __init__(self, name, customers_file="customers.xlsx"):
        self.name = name
        self.customers = []
        self.customers_file = customers_file

        # Загрузка данных о клиентах из файла Excel
        self.load_customers_data()

    def load_customers_data(self):
        try:
            workbook = load_workbook(self.customers_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, phone, loyalty_points = row
                customer = Customer(name, phone, loyalty_points)
                self.customers.append(customer)

            workbook.close()
        except FileNotFoundError:
            pass

    def save_customers_data(self):
        workbook = load_workbook(self.customers_file)
        sheet = workbook.active
        sheet.append(["Name", "Phone", "Loyalty Points"])

        for customer in self.customers:
            sheet.append([customer.name, customer.phone, customer.loyalty_points])

        workbook.save(self.customers_file)
        workbook.close()

    def add_customer(self, customer):
        self.customers.append(customer)

    def display_customer_services(self):
        workbook = load_workbook("beauty_salon_data.xlsx")
        sheet = workbook.active


        for row in sheet.iter_rows(min_row=1, values_only=True):
            print(f"Имя: {row[0]} Услуга: {row[1]} цена: {row[2]}")

        workbook.close()


    def display_customers(self):
        print(f"Клиенты салона '{self.name}':")
        for customer in self.customers:
            customer.display_info()

    def record_appointment(self, customer, service):
        workbook = load_workbook("beauty_salon_data.xlsx")
        sheet = workbook.active
        print(f"Клиент {customer.name} записан на услугу: {service}")
        customer.earn_loyalty_points(5)  # Например, начисление баллов за запись на услугу
        print(f"{customer.name} получил 5 баллов лояльности.")
        discounted_price = customer.apply_discount(service.price)
        print(f"С учетом скидки, клиент заплатит: ${discounted_price}")
        sheet.append([customer.name, service.name, discounted_price])
        workbook.save("beauty_salon_data.xlsx")

    def set_discount_rate(self, rate):
        Customer.set_discount_rate(rate)
        print(f"Скидка для всех клиентов салона '{self.name}' установлена на {rate * 100}%.")


def main():
    salon_name = input("Введите название салона красоты: ")
    beauty_salon = BeautySalon(salon_name)

    while True:
        print("\n1. Добавить клиента\n2. Показать клиентов\n3. Записать на услугу\n"
              "4. Установить скидку\n5. Показать клиентов и услуги\n6. Сохранить и выйти")
        choice = input("Введите номер действия: ")

        if choice == '1':
            name = input("Введите имя клиента: ")
            phone = input("Введите номер телефона клиента: ")
            customer = Customer(name, phone)
            beauty_salon.add_customer(customer)
            print(f"{name} добавлен в систему.")
            beauty_salon.save_customers_data()
        elif choice == '2':
            beauty_salon.display_customers()
        elif choice == '3':
            beauty_salon.display_customers()
            customer_name = input("Введите имя клиента: ")
            service_name = input("Введите название услуги: ")
            service_price = float(input("Введите цену услуги: $"))
            service = BeautyService(service_name, service_price)
            customer = next((c for c in beauty_salon.customers if c.name == customer_name), None)
            if customer:
                beauty_salon.record_appointment(customer, service)
            else:
                print(f"Клиент {customer_name} не найден.")
        elif choice == '4':
            discount_rate = float(input("Введите новую скидку в процентах: ")) / 100
            beauty_salon.set_discount_rate(discount_rate)
        elif choice == '5':
            beauty_salon.display_customer_services()
        elif choice == '6':
            # Сохранение данных о клиентах перед выходом
            beauty_salon.save_customers_data()
            print("Данные сохранены. До свидания!")
            break
        else:
            print("Некорректный ввод. Попробуйте еще раз.")


if __name__ == "__main__":
    main()
